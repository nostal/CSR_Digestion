from __future__ import division
from openpyxl.reader.excel import load_workbook
from random import shuffle
from copy import deepcopy

import psycopg2
import nltk
from nltk.tokenize import RegexpTokenizer, sent_tokenize
from gensim import corpora, models
from sklearn import svm
import pickle
import matplotlib.pyplot as plt
from sklearn import cross_validation
import numpy as np

# wb = load_workbook(filename=r"practice.xlsx")
# sheetnames = wb.get_sheet_names()
# ws = wb.get_sheet_by_name(sheetnames[0])
# data = []
# for i in range(ws.get_highest_column()):
#     for j in range(ws.get_highest_row()):
#         content = ws.cell(row=j, column=i).value
#         if content != None:
#             data.append(content)
#
# conn = psycopg2.connect("dbname=mydb user=postgres")
# cur = conn.cursor()
# cur.execute("select distinct description from products_practice;")
# more_data = cur.fetchall()
# more_data = [item[0] for item in more_data]
# data.extend(more_data)
#
# pattern = r"[a-zA-Z]+"
# tokenizer = RegexpTokenizer(pattern)
# data = [tokenizer.tokenize(item.lower()) for item in data]
#
# stopwords = []
# for line in open('stopwordslist.txt'):
#         stopwords.append(line.strip('\n'))
#
# wnl = nltk.WordNetLemmatizer()
# data = [[wnl.lemmatize(word) for word in document if word not in stopwords] for document in data]
#
# nonpractice_text = open('nonpractice.txt').read().lower()
# sents = sent_tokenize(nonpractice_text)
# nonpractice_documents = [tokenizer.tokenize(sent) for sent in sents]
# nonpractice_documents = [[wnl.lemmatize(word) for word in document if word not in stopwords] for document in nonpractice_documents]
#
# practice = deepcopy(data)
# nonpractice = nonpractice_documents
#
# data.extend(nonpractice_documents)
# allWords = sum(data, [])
#
# # freq = nltk.FreqDist(allWords)
# # keys = freq.keys()
# # wordFreq = open('Freq.txt', 'w')
# # for item in keys:
# #     wordFreq.write(item + '\n')
# # wordFreq.close()
# #
# # allWordsForNonPractice = sum(nonpractice_documents, [])
# # freq = nltk.FreqDist(allWordsForNonPractice)
# # keys = freq.keys()
# # wordFreq = open('FreqForNonPractice.txt', 'w')
# # for item in keys:
# #     wordFreq.write(item + '\n')
# # wordFreq.close()
#
#
#
# tokens_once = [word for word in set(allWords) if allWords.count(word) == 1]
# data = [[word for word in document if word not in tokens_once] for document in data]
#
# with open('data.pickle', 'wb') as f:
#     pickle.dump(data, f)
#
# with open('practice.pickle', 'wb') as f:
#     pickle.dump(practice, f)
#
# with open('nonpractice.pickle', 'wb') as f:
#     pickle.dump(nonpractice, f)

with open('data.pickle', 'rb') as f:
    data = pickle.load(f)

with open('practice.pickle', 'rb') as f:
    practice = pickle.load(f)

with open('nonpractice.pickle', 'rb') as f:
    nonpractice = pickle.load(f)

num = len(data)

dic = corpora.Dictionary.load('data.dict')
# dic = corpora.Dictionary(data)
# dic.save('data.dict')
#
corpus = [dic.doc2bow(document) for document in data]
def buildmodel(topicNum):
    model = models.LdaModel(corpus, id2word= dic, num_topics=topicNum, update_every=0, passes=20)
    return model

#    model.save('LDA')

# contents = model.show_topics(topics=10, topn=10, log=False, formatted=False)

# print contents
# keywords = open('keywords.txt', 'w')
# for topic in contents:
#     for word in topic:
#         keywords.write(word[1] + ' ')
#     keywords.write('\n')
# keywords.close()

def SVMTrain(model, topicNum):
    feature_set = np.zeros((num, topicNum))
    for i in range(num):
        feature_set[i] = [topic_dist[1] for topic_dist in model[corpus[i]]]

    # for item in corpus:
    #     feature = model[item]
    #     feature_set.append([topic_dist[1] for topic_dist in feature])
    # print feature_set

    label = np.zeros(num)
    label[0:len(practice)] = 1
    label[len(practice):] = 0

    SVM = svm.SVC()
    scores = cross_validation.cross_val_score(SVM, feature_set, label, cv=10)
    return scores.mean()

def main():
    scores = []
    for i in range(1,51):
        model = buildmodel(i)
        scores.append(SVMTrain(model, i))
    return scores

scores = main()

with open('scores.pickle', 'wb') as f:
    pickle.dump(scores, f)
